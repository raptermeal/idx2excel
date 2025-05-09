import requests
import csv
from datetime import datetime, timedelta
from io import StringIO
import os


# ✅ 날짜 설정
today = datetime.now()
start_date = today - timedelta(days=365)
end_date = today - timedelta(days=1)


# ✅ MarketWatch 지수 목록 (13개)
marketwatch_indices = {
    "USD/VND": "https://www.marketwatch.com/investing/currency/usdvnd/downloaddatapartial?startdate={}&enddate={}&daterange=d30&frequency=p1d&csvdownload=true&downloadpartial=false&newdates=false",
    "PSEi": "https://www.marketwatch.com/investing/index/psei/downloaddatapartial?startdate={}&enddate={}&daterange=d30&frequency=p1d&csvdownload=true&downloadpartial=false&newdates=false&countrycode=ph",
    "SEZE": "https://www.marketwatch.com/investing/index/szcp/downloaddatapartial?startdate={}&enddate={}&daterange=d30&frequency=p1d&csvdownload=true&downloadpartial=false&newdates=false&countrycode=xx",
    "DowJones": "https://www.marketwatch.com/investing/index/djia/downloaddatapartial?startdate={}&enddate={}&daterange=d30&frequency=p1d&csvdownload=true&downloadpartial=false&newdates=false",
    "S&P 500": "https://www.marketwatch.com/investing/index/spx/downloaddatapartial?startdate={}&enddate={}&daterange=d30&frequency=p1d&csvdownload=true&downloadpartial=false&newdates=false",
    "NASDAQ": "https://www.marketwatch.com/investing/index/comp/downloaddatapartial?startdate={}&enddate={}&daterange=d30&frequency=p1d&csvdownload=true&downloadpartial=false&newdates=false",
    "USD/CNY": "https://www.marketwatch.com/investing/currency/usdcny/downloaddatapartial?startdate={}&enddate={}&daterange=d30&frequency=p1d&csvdownload=true&downloadpartial=false&newdates=false",
    "USD/KRW": "https://www.marketwatch.com/investing/currency/usdkrw/downloaddatapartial?startdate={}&enddate={}&daterange=d30&frequency=p1d&csvdownload=true&downloadpartial=false&newdates=false",
    "KOSPI": "https://www.marketwatch.com/investing/index/180721/downloaddatapartial?startdate={}&enddate={}&daterange=d30&frequency=p1d&csvdownload=true&downloadpartial=false&newdates=false&countrycode=kr",
    "IDX": "https://www.marketwatch.com/investing/index/jakidx/downloaddatapartial?startdate={}&enddate={}&daterange=d30&frequency=p1d&csvdownload=true&downloadpartial=false&newdates=false&countrycode=id",
    "USD/IDR": "https://www.marketwatch.com/investing/currency/usdidr/downloaddatapartial?startdate={}&enddate={}&daterange=d30&frequency=p1d&csvdownload=true&downloadpartial=false&newdates=false",
    "USD/PHP": "https://www.marketwatch.com/investing/currency/usdphp/downloaddatapartial?startdate={}&enddate={}&daterange=d30&frequency=p1d&csvdownload=true&downloadpartial=false&newdates=false"
}


# ✅ 헤더와 쿠키 설정 (브라우저에서 가져온 최신 값 사용)
# ✅ 정교한 HTTP 헤더 설정
headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/136.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
    "Accept-Language": "ko,en-US;q=0.9,en;q=0.8,id;q=0.7",
    "Accept-Encoding": "gzip, deflate, br, zstd",
    "Connection": "keep-alive",
    "Referer": "https://www.marketwatch.com/",
    "Sec-Fetch-Dest": "document",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "none",
    "Sec-Fetch-User": "?1",
    "Upgrade-Insecure-Requests": "1"
}


# ✅ 쿠키 설정 (브라우저에서 복사한 쿠키 사용)
cookies = {
    "gdprApplies": "false",
    "letsGetMikey": "enabled",
    "mw_loc": '{"Region":"11","Country":"KR","Continent":"AS","ApplicablePrivacy":0}',
    "datadome": "88dXUmcFtvx19dkiQ80HLjAkkDafgx0rS4SKIo8GsjYtZxA8ucI5dGmu_VA6TREMNIVFHgkLO3qz9mwaPal4Vvh7mIjWNCN8XTX7EoqXd9iRcE2E47UluqOyWjVHw01D"
}


# ✅ 데이터 수집 함수
def fetch_marketwatch_data():
    combined_data = []
    session = requests.Session()
    session.headers.update(headers)
    session.cookies.update(cookies)


    for index_name, url_template in marketwatch_indices.items():
        url = url_template.format(start_date.strftime("%m/%d/%Y 00:00:00"), end_date.strftime("%m/%d/%Y 23:59:59"))
        response = session.get(url, timeout=10)
       
        if response.status_code == 200 and "text/csv" in response.headers.get("Content-Type", ""):
            reader = csv.reader(StringIO(response.text))
            next(reader)  # 헤더 건너뛰기
            for row in reader:
                if len(row) >= 5:
                    # ✅ 날짜 형식 변환: mm/dd/yyyy → yyyy-mm-dd
                    date_str = row[0].strip()
                    date_formatted = datetime.strptime(date_str, "%m/%d/%Y").strftime("%Y-%m-%d")
                   
                    combined_data.append({
                        "Index": index_name,
                        "Date": date_formatted,  # 변환된 날짜 형식 적용
                        # "Date": row[0],
                        "Open": row[1].replace(",", ""),
                        "High": row[2].replace(",", ""),
                        "Low": row[3].replace(",", ""),
                        "Close": row[4].replace(",", "")
                    })
            print(f"✅ {index_name} 데이터 다운로드 완료")
        else:
            print(f"❌ {index_name} 데이터 다운로드 실패: 상태 코드 {response.status_code} - {response.text[:200]}")
   
    return combined_data


# ✅ Naver 지수 목록 (VNI)
naver_indices = {
    "VNI": "https://api.stock.naver.com/index/.VNI/price?page={}&pageSize=50"
}


# ✅ Naver 데이터 수집 함수 (pandas 없이)
def fetch_naver_data():
    combined_data = []


    for page in range(1, 7):  # 페이지 1~6 (필요시 확장 가능)
        url = naver_indices["VNI"].format(page)
        response = requests.get(url, headers={"User-Agent": "Mozilla/5.0", "Accept": "application/json"})
       
        if response.status_code == 200:
            data = response.json()
            for item in data:
                combined_data.append({
                    "Date": item.get("localTradedAt", "").split("T")[0],
                    "Open": item.get("openPrice", "").replace(",", "").strip(),
                    "High": item.get("highPrice", "").replace(",", "").strip(),
                    "Low": item.get("lowPrice", "").replace(",", "").strip(),
                    "Close": item.get("closePrice", "").replace(",", "").strip(),
                    "Index": "VNI"
                })
            # print(f"✅ VNI 페이지 {page} 데이터 다운로드 완료")
        else:
            print(f"❌ VNI 페이지 {page} 데이터 다운로드 실패: 상태 코드 {response.status_code}")
    print(f"✅ VNI 데이터 다운로드 완료")
    return combined_data


# ✅ 데이터 수집 실행
df_marketwatch = fetch_marketwatch_data()
df_naver = fetch_naver_data()


# ✅ 데이터 병합 (pandas 없이)
df_combined = df_marketwatch + df_naver


# ✅ 국가, 구분, 단위, 출처 지정 (단일 딕셔너리)
index_info = {
    "USD/VND": {"국가": "베트남", "구분": "환율", "단위": "USD/VND", "출처": "MarketWatch"},
    "PSEi": {"국가": "필리핀", "구분": "주가", "단위": "PSEi", "출처": "MarketWatch"},
    "SEZE": {"국가": "중국", "구분": "주가", "단위": "SEZE", "출처": "MarketWatch"},
    "DowJones": {"국가": "미국", "구분": "주가", "단위": "DowJones", "출처": "MarketWatch"},
    "S&P 500": {"국가": "미국", "구분": "주가", "단위": "S&P 500", "출처": "MarketWatch"},
    "NASDAQ": {"국가": "미국", "구분": "주가", "단위": "NASDAQ", "출처": "MarketWatch"},
    "USD/CNY": {"국가": "중국", "구분": "환율", "단위": "USD/CNY", "출처": "MarketWatch"},
    "USD/KRW": {"국가": "한국", "구분": "환율", "단위": "USD/KRW", "출처": "MarketWatch"},
    "KOSPI": {"국가": "한국", "구분": "주가", "단위": "KOSPI", "출처": "MarketWatch"},
    "IDX": {"국가": "인도네시아", "구분": "주가", "단위": "IDX", "출처": "MarketWatch"},
    "USD/IDR": {"국가": "인도네시아", "구분": "환율", "단위": "USD/IDR", "출처": "MarketWatch"},
    "USD/PHP": {"국가": "필리핀", "구분": "환율", "단위": "USD/PHP", "출처": "MarketWatch"},
    "VNI": {"국가": "베트남", "구분": "주가", "단위": "VNI", "출처": "Naver"}
}


# ✅ df_combined 복사하여 새로운 리스트로 변환 (pandas 없이)
df_final = []


for row in df_combined:
    # ✅ 날짜 형식 자동 감지 및 변환
    date_str = row["Date"].strip()
    date_formatted = datetime.strptime(date_str, "%Y-%m-%d").strftime("%Y-%m-%d")
   
    # ✅ 국가, 구분, 단위, 출처 지정 (딕셔너리 기반)
    info = index_info.get(row["Index"], {})
    df_final.append({
        "인덱스": f"{row['Index']}_{date_formatted}",
        "국가": info.get("국가", ""),
        "구분": info.get("구분", ""),
        "단위": info.get("단위", ""),
        "날짜": date_formatted,
        "Open": float(row["Open"]) if row["Open"] else None,
        "High": float(row["High"]) if row["High"] else None,
        "Low": float(row["Low"]) if row["Low"] else None,
        "Close": float(row["Close"]) if row["Close"] else None,
        "출처": info.get("출처", "")
    })


# ✅ 열 순서 재정렬 (리스트에서 바로 지정)
df_final_sorted = [
    {
        "인덱스": row["인덱스"],
        "국가": row["국가"],
        "구분": row["구분"],
        "단위": row["단위"],
        "날짜": row["날짜"],
        "Open": row["Open"],
        "High": row["High"],
        "Low": row["Low"],
        "Close": row["Close"],
        "출처": row["출처"]
    }
    for row in df_final
]


# ✅ 결과 출력
print("\n✅ df_final 구조 변경 완료")
for row in df_final_sorted[:5]:
    print(row)
   
import shutil
from openpyxl import load_workbook


# ✅ 최종 결과 파일 저장 설정
template_path = "./data/idx_temp_250509.xlsx"  # 템플릿 파일 경로 지정
result_dir = "./result"
# ✅ 날짜 설정
today = datetime.today()
date_str = today.strftime("%y%m%d")
output_path = f"{result_dir}/keyidx_{date_str}.xlsx"     # 최종 파일 경로 지정
sheet_name = "table"            # 저장할 시트 이름
os.makedirs(result_dir, exist_ok=True)


# ✅ df_final_sorted 데이터 준비 (첫 번째 행: 헤더)
result_rows = [
    ["인덱스", "국가", "구분", "단위", "날짜", "Open", "High", "Low", "Close", "출처"]
] + [
    [
        row["인덱스"], row["국가"], row["구분"], row["단위"],
        row["날짜"], row["Open"], row["High"], row["Low"], row["Close"], row["출처"]
    ]
    for row in df_final_sorted
]


# ✅ 템플릿 복사 후 기록
shutil.copy(template_path, output_path)
wb = load_workbook(output_path)
if sheet_name not in wb.sheetnames:
    wb.create_sheet(sheet_name)
ws = wb[sheet_name]


# ✅ 기존 데이터 삭제
ws.delete_rows(1, ws.max_row)


# ✅ 새 데이터 작성 + 날짜 형식 지정
for row_idx, row in enumerate(result_rows, start=1):
    for col_idx, value in enumerate(row, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        # ✅ 날짜 형식 지정 (첫 번째 행 제외, 날짜 열만)
        if row_idx > 1 and col_idx == 5:  # 날짜열 (5번째 열)
            cell.number_format = "yyyy-mm-dd"


# ✅ 저장
wb.save(output_path)
print(f"\n📄 최종 저장 완료: {output_path} → 시트 '{sheet_name}' ({len(result_rows) - 1}행)")

